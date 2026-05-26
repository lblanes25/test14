"""
workbook_writer.py
==================
Turns the in-memory results from workbook_processor.process_workbook() into a
downloadable .xlsx, WITHOUT overwriting the original (Original Preservation).

Output layout:
  * Sheet 1: "_SUMMARY_" — the transformation workpaper. What changed, per
    sheet and operation, plus flags requiring review.
  * One sheet per cleaned table, keeping the original sheet name.
  * One hidden removed tab per sheet that had removed/flagged content, laid out
    as TWO labeled sections:
        REMOVED ROWS            (duplicates / blank rows — table schema)
        FLAGGED — OUTSIDE TABLE (title rows, notes, side tables — own shapes)
    The two are never merged into one ragged table.

Dependencies: pandas + numpy + openpyxl (all preinstalled in Code Interpreter).
"""

import os
import re
import math
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import cleanup_config as cfg


# ---------------------------------------------------------------------------
# Name + value helpers
# ---------------------------------------------------------------------------
_INVALID_SHEET = re.compile(r"[\\/?*\[\]:]")


def _safe_sheet_name(name, used):
    """Make an Excel-legal (<=31 char, no forbidden chars), unique sheet name."""
    s = _INVALID_SHEET.sub("_", str(name)).strip() or "Sheet"
    s = s[:31]
    base, i = s, 1
    while s.lower() in used:
        suffix = f"_{i}"
        s = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(s.lower())
    return s


def _clean_value(v):
    """openpyxl-safe scalar: turn NaN/NaT into None so Excel shows blanks."""
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def _write_df(ws, df, start_row, include_header=True):
    """Write a DataFrame to a worksheet starting at start_row (1-based).
    Returns the next free row after the block."""
    r = start_row
    if include_header:
        for c, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=r, column=c, value=_clean_value(col))
            cell.font = Font(bold=True)
        r += 1
    for _, row in df.iterrows():
        for c, val in enumerate(row.tolist(), start=1):
            ws.cell(row=r, column=c, value=_clean_value(val))
        r += 1
    return r


# ---------------------------------------------------------------------------
# Removed / flagged tab
# ---------------------------------------------------------------------------
_BANNER_FILL = PatternFill("solid", fgColor="FFE0E0")


def _write_removed_tab(wb, sheet_name, result, used_names):
    removed_rows = result.get("removed_rows", pd.DataFrame())
    flagged = result.get("flagged_regions", [])
    if (removed_rows is None or removed_rows.empty) and not flagged:
        return None

    tab_name = _safe_sheet_name(f"_RM_{sheet_name}", used_names)
    ws = wb.create_sheet(title=tab_name)
    ws.sheet_state = "hidden" if cfg.REMOVED_TAB_HIDDEN else "visible"

    # Banner
    banner = ws.cell(
        row=1, column=1,
        value=f"DO NOT ANALYZE — removed / flagged content from sheet "
              f"'{sheet_name}'. Originals preserved here for review.",
    )
    banner.font = Font(bold=True, color="990000")
    banner.fill = _BANNER_FILL
    r = 3

    # Section 1: removed rows (share the table schema)
    if removed_rows is not None and not removed_rows.empty:
        h = ws.cell(row=r, column=1, value="REMOVED ROWS (duplicates / blank rows)")
        h.font = Font(bold=True)
        r = _write_df(ws, removed_rows, r + 1)
        r += 2

    # Section 2: flagged out-of-table regions (each in its own shape)
    if flagged:
        h = ws.cell(row=r, column=1, value="FLAGGED — OUTSIDE MAIN TABLE BLOCK")
        h.font = Font(bold=True)
        r += 1
        for region in flagged:
            note = ws.cell(
                row=r, column=1,
                value=f"· {region['reason']} (origin row {region['origin_row']}, "
                      f"col {region['origin_col']})",
            )
            note.font = Font(italic=True)
            r += 1
            r = _write_df(ws, region["data"], r, include_header=False)
            r += 2
    return tab_name


# ---------------------------------------------------------------------------
# Summary sheet (the workpaper)
# ---------------------------------------------------------------------------
def _write_summary(wb, results):
    ws = wb.create_sheet(title="_SUMMARY_", index=0)
    title = ws.cell(row=1, column=1, value="Transformation Summary")
    title.font = Font(bold=True, size=14)
    r = 3
    for sheet_name, res in results.items():
        hdr = ws.cell(row=r, column=1, value=f"Sheet: {sheet_name}")
        hdr.font = Font(bold=True, size=12)
        r += 1
        ws.cell(row=r, column=1, value=res.get("note", "")); r += 1

        if res.get("cleaned") is None:
            r += 1
            continue

        for col, lbl in enumerate(["Operation", "Count", "Details"], start=1):
            c = ws.cell(row=r, column=col, value=lbl); c.font = Font(bold=True)
        r += 1
        for rec in res.get("summary", []):
            ws.cell(row=r, column=1, value=rec["operation"])
            ws.cell(row=r, column=2, value=rec["count"])
            ws.cell(row=r, column=3, value=rec["details"])
            r += 1

        flags = res.get("flags", [])
        if flags:
            fc = ws.cell(row=r, column=1, value="FLAGS — review before use:")
            fc.font = Font(bold=True, color="990000")
            r += 1
            for col, lbl in enumerate(["Condition", "Location", "Detail"], start=1):
                c = ws.cell(row=r, column=col, value=lbl); c.font = Font(bold=True)
            r += 1
            for f in flags:
                ws.cell(row=r, column=1, value=f["condition"])
                ws.cell(row=r, column=2, value=f["location"])
                ws.cell(row=r, column=3, value=f["detail"])
                r += 1
        r += 2

    # Widen columns a little for readability
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["C"].width = 80
    return ws


# ---------------------------------------------------------------------------
# Top-level
# ---------------------------------------------------------------------------
def write_workbook(results, original_path, output_path=None):
    """
    Build the cleaned workbook from process_workbook results.

    original_path : the source file — used only to derive a safe output name.
                    It is NEVER written to.
    output_path   : optional; defaults to '<original>_cleaned.xlsx'.
    Returns the output path.
    """
    if output_path is None:
        stem, _ = os.path.splitext(original_path)
        output_path = f"{stem}_cleaned.xlsx"
    if os.path.abspath(output_path) == os.path.abspath(original_path):
        raise ValueError("Refusing to overwrite the original file.")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    _write_summary(wb, results)

    used_names = {"_summary_"}
    for sheet_name, res in results.items():
        if res.get("cleaned") is None:
            continue
        name = _safe_sheet_name(sheet_name, used_names)
        ws = wb.create_sheet(title=name)
        _write_df(ws, res["cleaned"], 1)
        _write_removed_tab(wb, sheet_name, res, used_names)

    wb.save(output_path)
    return output_path
