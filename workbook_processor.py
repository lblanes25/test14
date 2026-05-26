"""
workbook_processor.py
=====================
Multi-tab orchestration layer on top of cleanup_engine.

Per the team's decisions:
  * Each sheet is scanned for the LARGEST contiguous block of populated cells.
    That block is treated as the table; everything outside it is FLAGGED
    (relocated), not cleaned. This is deterministic and conservative — we do
    not attempt to parse multiple informal tables per sheet (a v2 problem).
  * Removed/flagged content goes to ONE _REMOVED_ tab PER cleaned sheet.
  * Original workbook is never overwritten; output is a new file.

Dependencies: pandas + numpy + openpyxl (all preinstalled in Code Interpreter).

Block detection method:
  Split the sheet by fully-blank rows and fully-blank columns — the visual
  "gutters" people use to separate tables, titles, and notes. This yields a
  grid of candidate blocks; the one with the most populated cells wins.
"""

import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook

import cleanup_config as cfg
import cleanup_engine as ce
import validation_engine as ve


# ---------------------------------------------------------------------------
# Grid helpers
# ---------------------------------------------------------------------------
def _populated_mask(grid):
    """Boolean DataFrame: True where a cell holds a non-blank value."""
    def filled(v):
        if v is None:
            return False
        if isinstance(v, float) and np.isnan(v):
            return False
        return str(v).strip() != ""
    # Use per-column .map (applymap was removed in recent pandas).
    return grid.apply(lambda col: col.map(filled))


def _segments(flags):
    """
    Given a 1-D boolean array (True = keep / populated line), return list of
    (start, end_exclusive) runs of consecutive True. Used to find row-bands
    and column-bands separated by fully-blank lines.
    """
    segs, start = [], None
    for i, on in enumerate(flags):
        if on and start is None:
            start = i
        elif not on and start is not None:
            segs.append((start, i))
            start = None
    if start is not None:
        segs.append((start, len(flags)))
    return segs


# ---------------------------------------------------------------------------
# Table-block detection
# ---------------------------------------------------------------------------
def find_table_block(grid):
    """
    Find the largest contiguous populated block in a raw sheet grid
    (DataFrame read with header=None).

    Returns (r0, r1, c0, c1, blocks) where the bounds are inclusive-exclusive
    row/col indices of the winning block, and `blocks` is the full list of
    candidate (r0,r1,c0,c1,populated_count) for reporting.
    """
    mask = _populated_mask(grid)
    if mask.values.sum() == 0:
        return None  # empty sheet

    row_has = mask.any(axis=1).tolist()      # rows that aren't fully blank
    col_has = mask.any(axis=0).tolist()      # cols that aren't fully blank
    row_bands = _segments(row_has)
    col_bands = _segments(col_has)

    blocks = []
    for (r0, r1) in row_bands:
        for (c0, c1) in col_bands:
            count = int(mask.iloc[r0:r1, c0:c1].values.sum())
            if count > 0:
                blocks.append((r0, r1, c0, c1, count))

    # Winner = most populated cells; ties broken by topmost-then-leftmost
    # for determinism.
    winner = max(blocks, key=lambda b: (b[4], -b[0], -b[2]))
    r0, r1, c0, c1, _ = winner
    return r0, r1, c0, c1, blocks


# ---------------------------------------------------------------------------
# Date-column detection (conservative, content-based)
# ---------------------------------------------------------------------------
_DATEISH = re.compile(r"^\s*\d{1,4}[/\-.]\d{1,2}[/\-.]\d{1,4}\s*$")


def detect_date_columns(df, min_fraction=0.7):
    """
    Conservatively identify columns to attempt date normalization on:
    a column qualifies only if >= min_fraction of its non-null values look
    like 3-part dates. Header-name guessing is intentionally NOT used alone
    (too easy to mislabel); content is the signal.
    """
    cols = []
    for col in df.columns:
        vals = df[col].dropna().astype(str)
        if len(vals) == 0:
            continue
        hits = vals.map(lambda v: bool(_DATEISH.match(v))).mean()
        if hits >= min_fraction:
            cols.append(col)
    return cols


# ---------------------------------------------------------------------------
# Per-sheet processing
# ---------------------------------------------------------------------------
def process_sheet(grid, sheet_name, ws=None):
    """
    Process one sheet given its raw grid (header=None DataFrame).
    ws: optional openpyxl worksheet for structural checks (merged/hidden).
    Returns a dict with the cleaned df, summary, removed frame, flags.
    """
    block = find_table_block(grid)
    if block is None:
        return {
            "sheet": sheet_name, "cleaned": None, "summary": [],
            "removed_rows": pd.DataFrame(), "flagged_regions": [],
            "date_flags": [],
            "note": "Sheet empty or no populated cells; skipped.",
        }

    r0, r1, c0, c1, blocks = block

    # Out-of-table content = populated cells outside the winning block.
    flagged_cells = []
    mask = _populated_mask(grid)
    for (br0, br1, bc0, bc1, cnt) in blocks:
        if (br0, br1, bc0, bc1) == (r0, r1, c0, c1):
            continue
        sub = grid.iloc[br0:br1, bc0:bc1]
        flagged_cells.append((br0, bc0, sub))

    # Extract the table; first row of the block is the header.
    table = grid.iloc[r0:r1, c0:c1].reset_index(drop=True)
    header = table.iloc[0].tolist()
    table = table.iloc[1:].reset_index(drop=True)
    table.columns = header

    date_cols = detect_date_columns(table)
    cleaned, summary, removed_frames, date_flags, numeric_flags = ce.run_cleanup(
        table, date_columns=date_cols
    )

    # Validation / flagging (describe-only, never modifies data).
    # Structural checks run on the raw worksheet BEFORE flattening; data-quality
    # checks run on the detected table; engine outcomes are merged in too.
    structural_flags = ve.validate_worksheet(ws) if ws is not None else []
    table_flags = ve.validate_table(table)
    engine_flags = ve.merge_engine_flags(date_flags, numeric_flags)
    all_flags = structural_flags + table_flags + engine_flags

    # Removed ROWS (duplicates, blank rows) share the table's schema — safe to
    # stack together. Keep them SEPARATE from out-of-table regions, which have
    # unrelated shapes; concatenating the two produces a ragged mess.
    removed_rows = (
        pd.concat(removed_frames, ignore_index=True)
        if removed_frames else pd.DataFrame()
    )

    # Flagged out-of-table regions: each kept in its own original shape, with
    # an origin note. The Excel writer lays these out as a labeled section
    # below the removed rows — never merged into one table.
    flagged_regions = []
    for (br0, bc0, sub) in flagged_cells:
        flagged_regions.append({
            "origin_row": br0 + 1,
            "origin_col": bc0 + 1,
            "reason": "outside main table block",
            "data": sub.reset_index(drop=True),
        })

    return {
        "sheet": sheet_name,
        "cleaned": cleaned,
        "summary": summary,
        "removed_rows": removed_rows,
        "flagged_regions": flagged_regions,
        "date_flags": date_flags,
        "numeric_flags": numeric_flags,
        "flags": all_flags,
        "detected_date_columns": date_cols,
        "block_bounds": (r0, r1, c0, c1),
        "note": (
            f"Table detected at rows {r0 + 1}-{r1}, cols {c0 + 1}-{c1}. "
            f"{len(flagged_regions)} out-of-table region(s) flagged."
        ),
    }


# ---------------------------------------------------------------------------
# Workbook orchestrator
# ---------------------------------------------------------------------------
def process_workbook(path):
    """
    Read every sheet, process each, and return a per-sheet result dict.
    Reading with header=None preserves the raw grid so block detection can
    see title rows and stray cells.
    """
    sheets = pd.read_excel(path, sheet_name=None, header=None, dtype=object)
    # Also open with openpyxl for structural checks (merged/hidden cells) that
    # pandas discards. data_only=True so formula cells read as their values.
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        wb = None
    results = {}
    for name, grid in sheets.items():
        ws = wb[name] if (wb is not None and name in wb.sheetnames) else None
        results[name] = process_sheet(grid, name, ws=ws)
    return results
